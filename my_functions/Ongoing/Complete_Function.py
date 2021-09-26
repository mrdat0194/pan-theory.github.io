#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue May 28 08:42:24 2019

@author: petern
"""
import datetime
import numpy as np
import time
import shutil
import pandas as pd
from main_def import MAIN_DIR

import os
import pickle
import bcrypt

def get_config():
    """
        How to enscrypt a strings, just for fun here.
        >>> import bcrypt
        >>> password = b"super secret password"
        >>> # Hash a password for the first time, with a randomly-generated salt
        >>> hashed = bcrypt.hashpw(password, bcrypt.gensalt())
        >>> # Check that an unhashed password matches one that has previously been
        >>> # hashed
        >>> if bcrypt.checkpw(password, hashed):
            ...     print("It Matches!")
        ... else:
        ...     print("It Does not Match :(")

        # abc = get_config()
        # print(abc)
    """

    path = os.path.join(MAIN_DIR, "ggl_api")
    path_save = os.path.join(path, "users_config.pkl")

    try:
        # see if we have run  this before
        with  open(path_save, 'rb') as users:
            users_dict = pickle.load(users)

    except IOError:
        # if not set to default
        print(IOError)

    username = input("Enter your username: ")
    password = input("Enter your password: ")
    get_what = input("What you wanna get: ")
    password = bytes(password, 'utf-8')
    hashed = users_dict.get(username)

    if bcrypt.checkpw(password, hashed): # unless a password can be None we can use get
        with open(path_save, 'rb') as users:
            users_dict = pickle.load(users)
            return users_dict.get(get_what)
    else:
        print("Please let it go!")

class list_edit:
    def merge_to_list(*lists):
        '''
        Merge and sort list
        :param lists: <list1>,<list2>
        :return: a sorted list
        Eg: list1 = ["a","b","c"]
            list2 = ["a", "d", "f"]
            list 3 = merge_to_list(list1,list2)
        '''
        print(lists)
        print(type(list))
        newlist = []
        for i in lists:
            newlist.extend(i)
        merge_list = set(newlist)
        merge_list = list(merge_list)
        merge_list.sort()
        return merge_list

class Move_file:
    """
    Move file with ease
    """

    def Data_move(Find_name , Path , Destination_folder):
        """
        Path = "/Users/petern/"
        Destination_folder = "/Users/petern/Desktop/GonJoy/Right Time/OnData/Tool_chat/"
        Find_name = "mytext"
        Move_file.Data_move(Find_name,Find_name, Path, Destination_folder)
        """
        entries = os.listdir(Path)
        for i in entries:
            if pd.Series(i).str.contains(Find_name).bool():
                path = Path + i
                Destination_folder = Destination_folder + i
                shutil.move(path, Destination_folder)

class Time_converse:
    """
    Process time series
    """
    def Time_diff(Segmentation1, id_time, time_col):
        """

        :param Segmentation1:
        :param id_time:
        :param time_col:
        :return:
        """
        Segmentation1['Timediff'] = float()
        Segmentation1[time_col] = pd.to_datetime(Segmentation1[time_col])        
        for i in range(1,Segmentation1.shape[0]):
            if ~pd.isna(Segmentation1[id_time][i]):
                if (Segmentation1[id_time][i] == Segmentation1[id_time][i-1] 
                and Segmentation1[time_col][i].month == Segmentation1[time_col][i-1].month 
                and Segmentation1[time_col][i].day == Segmentation1[time_col][i-1].day 
                and Segmentation1[time_col][i].hour == Segmentation1[time_col][i-1].hour):
                    Segmentation1["Timediff"][i] = (Segmentation1[time_col][i] - Segmentation1[time_col][i-1]).seconds
                elif (Segmentation1[time_col][i].month == Segmentation1[time_col][i-1].month 
                      and Segmentation1[time_col][i].day == Segmentation1[time_col][i-1].day 
                      and Segmentation1[time_col][i].hour == Segmentation1[time_col][i-1].hour):
                    Segmentation1["Timediff"][i-1] = (Segmentation1[time_col][i] - Segmentation1[time_col][i-1]).seconds
                    Segmentation1["Timediff"][i] = (Segmentation1[time_col][i] - Segmentation1[time_col][i-1] ).seconds
                    if (Segmentation1["Timediff"][i-1] == Segmentation1["Timediff"][i]):
                        Segmentation1["Timediff"][i-1] = 0
                else:
                    Segmentation1["Timediff"][i] = 0
            
        return Segmentation1


class Chat_analyse:
    """
        # Connect_final2 = Chat_analyse.teamcheck(Connect_final1, "from_id", "KM_QR", "KM_Content",
        #                                         "khuyen mai|km|khuyến mãi", "ukm|ikm|dkm|kmxpm", Connect_final1)
      Sequence of chat data and create column
    """
    def teamcheck(self, from_id ,AN1_QR, AN1_Content, ANsay, AN_not, Connect_final):
          col1 = [c for c in Connect_final.columns if pd.Series(c).str.contains(from_id).bool()]          
          Connect_final.rename(columns = {col1[0] : "from_id"}, inplace=True)          
          Connect_final =  Connect_final.reset_index(drop = True)
          Connect_final[AN1_QR] =  0
          Connect_final[AN1_Content] =  "0"
          Col_change1 = [c for c in Connect_final.columns if pd.Series(c).str.contains(AN1_QR).bool()]
          Col_change2 = [c for c in Connect_final.columns if pd.Series(c).str.contains(AN1_Content).bool()]          
          MonMansay = ANsay.lower()
          MonMan_not  = AN_not.lower()   
          Connect_final['message'] = Connect_final['message'].fillna("0")                  
          col_mes = [c for c in Connect_final.columns if pd.Series(c).str.contains('message').bool()]
          for k in range(len(Connect_final['from_id'])):          
              An_num = [c for c in Connect_final.loc[k, col_mes].astype(str) if pd.Series(c).str.lower().str.contains(MonMansay).bool()]
              if not An_num == []:  
                  Connect_final.loc[k, Col_change1] = 1
                  Connect_final.loc[k, Col_change2] = Connect_final.loc[k, col_mes].values
          for k in range(len(Connect_final['from_id'])):
              AN_not_num = [c for c in Connect_final.loc[k, col_mes].astype(str) if pd.Series(c).str.lower().str.contains(MonMan_not).bool()]
              if not AN_not_num == []:  
                  Connect_final.loc[k, Col_change1] = 0
                  Connect_final.loc[k, Col_change2] = 0                 
          col_t = [c for c in Connect_final.columns if pd.Series(c).str.contains('from_id').bool()]      
          Connect_final.rename(columns = {col_t[0] : from_id}, 
                     inplace=True)
          return Connect_final


         
def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
        # append_df_to_excel("hello.xlsx", pd.DataFrame(np.array([[1, 2, 3], [4, 5, 6], [7, 8, 9]]),
        #                columns=['a', 'b', 'c']))
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None

    """
    from openpyxl import load_workbook

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()

def is_prime(x):
    '''

    the original prime functions that could be improved.
        prime_or_not = is_prime(991)
        print(prime_or_not)
    :param x:
    :return:
    '''
    for i in range(2, int(x**0.5) + 1):
        if x%i == 0:
            return False
        else:
            print(x)
            return True

# Python program to display all the prime numbers within an interval
def prime_interval(x):
    lower = 900
    upper = 1000
    print("Prime numbers between", lower, "and", upper, "are:")

    for num in range(lower, upper + 1):
        # all prime numbers are greater than 1
        if num > 1:
            for i in range(2, num):
                if (num % i) == 0:
                    break
            else:
                print(num)

# Python program to print
# print all primes in a range
# using concept of Segmented Sieve
from math import floor, sqrt

# This functions finds
# all primes smaller than limit
# using simple sieve of eratosthenes.
# It stores found
# primes in list prime[]


def simpleSieve(limit, primes):
    mark = [False]*(limit+1)

    for i in range(2, limit+1):
        if not mark[i]:

            # If not marked yet,
            # then its a prime
            primes.append(i)
            for j in range(i, limit+1, i):
                mark[j] = True


# Finds all prime numbers
# in given range using
# segmented sieve
def primesInRange(low, high):
    """
        # Driver code
        # low = 10
        # high = 100
        # primesInRange(low, high)
    :param low:
    :param high:
    :return:
    """

    # Comput all primes smaller
    # or equal to
    # square root of high
    # using simple sieve
    limit = floor(sqrt(high)) + 1
    primes = list()
    simpleSieve(limit, primes)

    # Count of elements in given range
    n = high - low + 1

    # Declaring boolean only for
    # [low, high]
    mark = [False]*(n+1)

    # Use the found primes by
    # simpleSieve() to find
    # primes in given range
    for i in range(len(primes)):

        # Find the minimum number
        # in [low..high] that is
        # a multiple of prime[i]
        # (divisible by prime[i])
        loLim = floor(low/primes[i]) * primes[i]
        if loLim < low:
            loLim += primes[i]
        if loLim == primes[i]:
            loLim += primes[i]

        # Mark multiples of primes[i]
        # in [low..high]:
        # We are marking j - low for j,
        # i.e. each number
        # in range [low, high] is mapped
        # to [0, high-low]
        # so if range is [50, 100]
        # marking 50 corresponds
        # to marking 0, marking 51
        # corresponds to 1 and
        # so on. In this way we need
        # to allocate space
        # only for range
        for j in range(loLim, high+1, primes[i]):
            mark[j-low] = True

    # Numbers which are not marked
    # in range, are prime
    for i in range(low, high+1):
        if not mark[i-low]:
            print(i, end=" ")


def twoSumHashing(num_arr, pair_sum):
    sums = []
    hashTable = {}

    for i in range(len(num_arr)):
        complement = pair_sum - num_arr[i]
        if complement in hashTable:
            print("Pair with sum", pair_sum,"is: (", num_arr[i],",",complement,")")
        hashTable[num_arr[i]] = num_arr[i]

def my_fib(x, memo=dict()):
    if memo.get(x):
        return memo[x]
    if x == 1 or x == 2:
        result = 1
    else:
        result = my_fib(x - 1, memo) + my_fib(x -2, memo)
    memo[x] = result
    return result

def fibo(N):
    a = b = 1
    for _ in range(2,N): a,b = b,a+b
    return b

def simpleFibo(N,a=0,b=1):
    if N < 3: return a+b
    return simpleFibo(N-1,b,a+b)

def dynaFibo(N,memo={1:1,2:1}):
    '''Dynamic Programming'''
    if N not in memo:
        memo[N] = dynaFibo(N-1,memo) + dynaFibo(N-2,memo)
    return memo[N]

def dynaFibo2(N,memo=None):
    '''Dynamic Programming'''
    if not memo:    memo = [0,1,1]+[0]*N
    if not memo[N]: memo[N] = dynaFibo2(N-1,memo) + dynaFibo2(N-2,memo)
    return memo[N]

cache = [0, 1]   # Initialize with the first two terms of Fibonacci series.
def fibonacci(n):
    """
    Returns the n-th number in the Fibonacci sequence.
    bottom up
    Parameters
    ----------
    n: int
       The n-th number in the Fibonacci sequence.
    """
    for i in range(2, n):
        cache.append(cache[i-1] + cache[i-2])
    return cache[-1]

def binFibo(N):
    a,b   = 0,1
    f0,f1 = 1,1
    r,s   = (1,1) if N&1 else (0,1)
    N //=2
    while N > 0:
        a,b   = f0*a+f1*b, f0*b+f1*(a+b)
        f0,f1 = b-a,a
        if N&1: r,s = f0*r+f1*s, f0*s+f1*(r+s)
        N //= 2
    return r

def test_fib():
    from timeit import timeit
    count = 100

    N = 990

    t= timeit(lambda:my_fib(N,dict()), number=count) # providing dict() to avoid reuse between repetitions
    print("my_fib(N)",t)

    t= timeit(lambda:fibo(N), number=count)
    print("fibo(N)",t)

    t= timeit(lambda:simpleFibo(N), number=count)
    print("simpleFibo(N)",t)

    t= timeit(lambda:dynaFibo(N,{1:1,2:1}), number=count) # providing dict() to avoid reuse between repetitions
    print("dynaFibo(N)",t)

    t= timeit(lambda:dynaFibo2(N), number=count)
    print("dynaFibo2(N)",t)

    t= timeit(lambda:binFibo(N), number=count)
    print("binFibo(N)",t)

if __name__ == "__main__":
    pd.set_option('float_format', '{:,.2f}'.format)
    pd.set_option("display.max_rows", None, "display.max_columns", 60, 'display.width', 1000)


    # data_trial = os.path.join(MAIN_DIR, "data", "Gonj/")
    #
    # gonjoybot_list = sorted( os.listdir(data_trial), reverse = True )[0:2]
    #
    # li = []
    #
    # for filename in gonjoybot_list:
    #     filename = data_trial + filename
    #     df = pd.read_csv(filename, index_col=None, header=0)
    #     li.append(df)
    #
    # gonjoybot_chat = pd.concat(li, axis=0, ignore_index=True)
    #
    #
    # gonjoybot_chat['Date'] = pd.to_datetime(gonjoybot_chat['time']) +   pd.Timedelta(hours=7)
    #
    # gonjoybot_chat = gonjoybot_chat[gonjoybot_chat['Date'].dt.date == pd.to_datetime("2019-06-25")]
    #
    # #
    #
    # col1 = [c for c in gonjoybot_chat.columns if pd.Series(c).str.contains('^from|^time$').bool()]
    #
    # Segmentation = gonjoybot_chat[col1]
    #
    # Segmentation1 = Segmentation[~pd.isna(Segmentation["from_id"]) ].reset_index()
    #
    # Segmentation2 = Time_converse.Time_diff(Segmentation1, "from_id", "time")
    #
    # print(Segmentation2)
    #
    # Segmentation3 = Segmentation2.sort_values(by=['time'])
    #
    # print(Segmentation3)

    # # Driver Code for 2 sum
    # num_arr = [4, 5, 1, 8]
    # pair_sum = 9
    #
    # # Calling function
    # twoSumHashing(num_arr, pair_sum)


    ## Not fixed yet
    #Extract_Diff_id <- function(Event.Category, Var1, tota_FB_promotion,Question_user) {
    #
    #    col3 <- grep(Var1,colnames(Question_user))
    #
    #    names(Question_user)[col3] <- paste("Event.Category")
    #
    #    #
    #
    #    col2 <- grep(Event.Category,colnames(tota_FB_promotion))
    #
    #    names(tota_FB_promotion)[col2] <- paste("Event.Category")
    #
    #
    #
    #    col1 <- grep("Event.Category",colnames(Question_user))
    #
    #    for (i in 1:dim(Question_user)[1]) {
    #      if (i==1) {
    #        tota_FB_promotion <- tota_FB_promotion[which(as.character(tota_FB_promotion$Event.Category) != as.character(Question_user[i,col1])),]
    #      } else {
    #        tota_FB_promotion <- tota_FB_promotion[which(as.character(tota_FB_promotion$Event.Category) != as.character(Question_user[i,col1])),]
    #      }
    #    }
    #
    #    names(tota_FB_promotion)[col2] <- paste(Event.Category)
    #
    #    return  (tota_FB_promotion)
    #
    #
    #}
    #
    #
    ## Same extract
    ### Trich ID:
    #
    #Same_extract_id <- function(abc_b, abc_c,gonjoybot_chat, Chat_phone_usertx3) {
    #
    #  col3 <- grep(abc_c,colnames(Chat_phone_usertx3))
    #
    #  names(Chat_phone_usertx3)[col3] <- paste("Event Category")
    #
    #  #
    #
    #  col2 <- grep(abc_b,colnames(gonjoybot_chat))
    #
    #  names(gonjoybot_chat)[col2] <- paste("Event Category")
    #
    #  #
    #
    #
    #  for (i in 1:dim(Chat_phone_usertx3)[1]) {
    #    if (i==1) {
    #      Chat_phone_usert3_1 <- gonjoybot_chat[which(as.character(gonjoybot_chat$`Event Category`) == as.character(Chat_phone_usertx3$`Event Category`)[i]),]
    #    } else {
    #      Chat_phone_usert3_2 <- gonjoybot_chat[which(as.character(gonjoybot_chat$`Event Category`) == as.character(Chat_phone_usertx3$`Event Category`)[i]),]
    #      Chat_phone_usert3_1 <- rbind(Chat_phone_usert3_1,Chat_phone_usert3_2)
    #    }
    #  }
    #
    #  names(Chat_phone_usert3_1)[col2] <- paste(abc_b)
    #
    #  return (Chat_phone_usert3_1)
    #
    #}
    #def Get_id_GA_full(Joy_GA,No,DateUpdate):
    #
    #
    #  setwd("~/Desktop/GonJoy/Right Time/")
    #
    #  DateUpdate <- DateUpdate
    #
    #  setwd("~/Desktop/GonJoy/Right Time/OnData/Datarequire/")
    #
    #  DatabaseRequire_AllIn <- do.call(rbind,lapply(list.files(path = "~/Desktop/GonJoy/Right Time/OnData/Datarequire/"), read_csv))
    #
    #  DatabaseRequire_AllIn$date <- DatabaseRequire_AllIn$Date
    #
    #  DatabaseRequire_AllIn$date <- substr(DatabaseRequire_AllIn$date,1,8)
    #
    #  DatabaseRequire_AllIn$date <- anydate(  DatabaseRequire_AllIn$date)
    #
    #  DatabaseRequire_AllIn$date <- anydate( DatabaseRequire_AllIn$date)
    #
    #  DatabaseRequire_AllIn <- DatabaseRequire_AllIn[which(as.Date(DatabaseRequire_AllIn$date) >= DateUpdate ), ]
    #
    #  for (i in 1:length(DatabaseRequire_AllIn$`Event Category`)) {
    #
    #    start_string <- str_locate_all(DatabaseRequire_AllIn[i,4],"-")
    #
    #    start_string <- data.frame(start_string)
    #
    #    DatabaseRequire_AllIn[i,3] <- substr(DatabaseRequire_AllIn[i,4], start_string[2,2]+2,
    #                                         as.integer(nchar(as.character(DatabaseRequire_AllIn[i,4]))))
    #
    #    start_string <- str_locate_all(DatabaseRequire_AllIn[i,3],"-")
    #
    #    if (!isempty(start_string[[1]])) {
    #
    #      start_string <- data.frame(start_string)
    #
    #      DatabaseRequire_AllIn[i,3] <- substr(DatabaseRequire_AllIn[i,3], start_string[1,2]+2,
    #                                           as.integer(nchar(as.character(DatabaseRequire_AllIn[i,3]))))
    #    }
    #
    #  }
    #
    #  DatabaseRequire_AllIn1 <- DatabaseRequire_AllIn[,grep("Event label|Event Category", names(DatabaseRequire_AllIn))]
    #
    #  names(DatabaseRequire_AllIn1)[1] <- paste("Var1")
    #
    #  names(Joy_GA)[No] <- paste("Var1")
    #
    #  Joy_GA1 <- merge(Joy_GA,DatabaseRequire_AllIn1, by = "Var1", all.x = T )
    #
    #  Joy_GA2 <- unique(Joy_GA1)
    #
    #  return (Joy_GA2)
    #
    #}

    #
    ##class Database:
    ##
    ##    def Database_Loyalty(loyaltyCustomer, loyaltyProgram, Customer, Provider)
    ##
    ##    Database_Loyalty <- function (loyaltyCustomer, loyaltyProgram, Customer, Provider) {
    ##
    ##  # setwd("~/Desktop/GonJoy/Right Time/OnData/Datatest/Loyalty")
    ##  #
    ##  # loyaltyCustomer <- read.csv("~/Desktop/GonJoy/Right Time/OnData/Datatest/Loyalty/gonjoy34.loyaltyCustomers.csv")
    ##  #
    ##  # loyaltyProgram <- read.csv("~/Desktop/GonJoy/Right Time/OnData/Datatest/Loyalty/gonjoy34.loyaltyPrograms.csv")
    ##  #
    ##  # Customer <- read.csv("~/Desktop/GonJoy/Right Time/OnData/Datatest/Customers/gonjoy34.customer.csv")
    ##  #
    ##  # Provider <- read.csv("~/Desktop/GonJoy/Right Time/OnData/Datatest/Provider/gonjoy34.providers.csv")
    ##  #
    ##  col1 <- grep("X_id", names(Customer))
    ##
    ##  col2 <- grep("X_id", names(Provider))
    ##
    ##  col3 <- grep("createdAt", names(Customer))
    ##
    ##  col4 <- grep("createdAt", names(Provider))
    ##
    ##  names(Customer)[col1] <- paste("customer")
    ##
    ##  names(Provider)[col2] <- paste("provider")
    ##
    ##  names(Customer)[col3] <- paste("createdAt_Cus")
    ##
    ##  names(Provider)[col4] <- paste("createdAt_Pro")
    ##
    ##  #Col_Provider
    ##
    ##  cul_Customer1 <- grep("provider", colnames(Provider))
    ##  cul_Customer2 <- grep("name", colnames(Provider))[1]
    ##  cul_Customer3 <- grep("synonyms", colnames(Provider))
    ##  cul_Customer4 <- grep("gps_coordinates", colnames(Provider))[1]
    ##  cul_Customer6 <- grep("place", colnames(Provider))[1]
    ##
    ##  #Col_cust
    ##
    ##  col_Customer1 <- grep("fbId", colnames(Customer))
    ##  col_Customer2 <- grep("name", colnames(Customer))[1]
    ##  col_Customer3 <- grep("createdAt", colnames(Customer))
    ##  col_Customer4 <- grep("lastActiveTime", colnames(Customer))
    ##  col_Customer5 <- grep("customer", colnames(Customer))
    ##  col_Customer6 <- grep("X_id", colnames(Customer))
    ##
    ##  Customer1 <- Customer[,c(col_Customer1,col_Customer2,col_Customer3, col_Customer4, col_Customer5, col_Customer6)]
    ##
    ##  Provider1 <- Provider[,c(cul_Customer1,cul_Customer2,cul_Customer3, cul_Customer4, cul_Customer6)]
    ##
    ##  Loyalty_customer <- merge(loyaltyCustomer, Customer1, by = "customer", all.x = T)
    ##
    ##  Loyalty_customer$fbId <- as.character(Loyalty_customer$fbId)
    ##
    ##  Loyalty_customer1 <- Loyalty_customer
    ##
    ##  # Loyalty program - customer - provider
    ##
    ##  Loyalty_Provider <- merge(Loyalty_customer1, Provider1, by = "provider", all.x = T)
    ##
    ##  Loyalty_Provider1 <- Loyalty_Provider[!duplicated(Loyalty_Provider$fbId),]
    ##
    ##  co_Customer0 <- grep("X_id", colnames(loyaltyProgram))[1]
    ##  co_Customer1 <- grep("name", colnames(loyaltyProgram))[1]
    ##  co_Customer2 <- grep("provider", colnames(loyaltyProgram))[1]
    ##  co_Customer3 <- grep("description", colnames(loyaltyProgram))
    ##  co_Customer4 <- grep("levels", colnames(loyaltyProgram))[1]
    ##  co_Customer5 <- grep("updatedAt", colnames(loyaltyProgram))[1]
    ##
    ##  loyaltyProgram1 <- loyaltyProgram[,c(co_Customer0,co_Customer1,co_Customer2,co_Customer3,co_Customer4,co_Customer5)]
    ##
    ##  names(loyaltyProgram1)[6] <- paste("updatedAt_program")
    ##
    ##  #
    ##
    ##  Loyalty_Provider_Customer_Program <- merge(Loyalty_Provider, loyaltyProgram1, by = "provider", all.x = T)
    ##
    ##  return(Loyalty_Provider_Customer_Program)
    ##
    ##}